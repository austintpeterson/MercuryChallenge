from twitterscraper import query_tweets
import random
import requests
import datetime as dt
import datetime
import json
from functools import partial
from multiprocessing.pool import Pool
import csv
import collections
import openpyxl
import xlsxwriter
from nltk.tokenize import word_tokenize
import string
from nltk.corpus import stopwords
import operator



from twitterscraper.tweet import Tweet
from twitterscraper.ts_logger import logger
import urllib

class wCount:
    def __init__(self, word, count):
        self.word = word
        self.count = count

class JSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if hasattr(obj, '__json__'):
            return obj.__json__()
        elif isinstance(obj, collections.Iterable):
            return list(obj)
        elif isinstance(obj, dt.datetime):
            return obj.isoformat()
        elif hasattr(obj, '__getitem__') and hasattr(obj, 'keys'):
            return dict(obj)
        elif hasattr(obj, '__dict__'):
            return {member: getattr(obj, member)
                    for member in dir(obj)
                    if not member.startswith('_') and
                    not hasattr(getattr(obj, member), '__call__')}

        return json.JSONEncoder.default(self, obj)

def valid_date(s):
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        msg = "Not a valid date: '{0}'.".format(s)
        raise argparse.ArgumentTypeError(msg)

if __name__ == '__main__':
    """
	query_tweets(query, limit=None, begindate=dt.date(2006,3,21), enddate=dt.date.today(), poolsize=20, lang='')

	list_of_tweets = query_tweets("Shake Shack", 10, dt.date(2010,3,21), dt.date(2011,3,21), 20, 'ar')

	List of arugments and a sample function call 
    """
    kw = []
    dates = []
    ids = []
    
    wb = openpyxl.load_workbook('Funeral Keywords.xlsx')	
    sheet = wb.active
    col = sheet['B']
    
    wb2 = openpyxl.load_workbook('db.xlsx')	
    sheet2 = wb2.active
    col1 = sheet2['A']
    col2 = sheet2['B']
    for x in range(1, 17):
       kw.append(col[x].value)

    
    for y in range(0, len(col1)):
        dates.append(datetime.date.fromisoformat(col1[y].value))
        ids.append(col2[y].value)
        
    
    # Create an new Excel file and convert the text data.
    workbook = xlsxwriter.Workbook("funeraloutput.xlsx")
    worksheet = workbook.add_worksheet()
    

    # Start from the first cell.
    row = 0
    num = 1
    worksheet.write(row, 0, "Number")
    worksheet.write(row, 1, "user")
    worksheet.write(row, 2, "fullname")
    worksheet.write(row, 3, "tweet-id")
    worksheet.write(row, 4, "timestamp")
    worksheet.write(row, 5, "url")
    worksheet.write(row, 6, "likes")
    worksheet.write(row, 7, "replies")
    worksheet.write(row, 8, "retweets")
    worksheet.write(row, 9, "text")
    worksheet.write(row, 10, "html")
    worksheet.write(row, 11, "event id")
    worksheet.write(row, 12, "event date")

    row += 1

    for d in range(0, len(dates)): 
    
        if dates[d].month == 3 and dates[d].year == 2017:
            mindate = dates[d] - datetime.timedelta(days=3)
            maxdate = dates[d] + datetime.timedelta(days=7)
            
            for y in range(0, len(kw)):
     
                list_of_tweets = query_tweets(kw[y], 100, mindate, maxdate, 20)
            
            
                # Read the text file and write it to the worksheet.
                for x in list_of_tweets:

                # Write any other lines to the worksheet.
                    worksheet.write(row, 0, num)
                    worksheet.write(row, 1, x.user)
                    worksheet.write(row, 2, x.fullname)
                    worksheet.write(row, 3, x.id)
                    worksheet.write(row, 4, x.timestamp)
                    worksheet.write(row, 5, x.url)
                    worksheet.write(row, 6, x.likes)
                    worksheet.write(row, 7, x.replies)
                    worksheet.write(row, 8, x.retweets)
                    worksheet.write(row, 9, x.text)
                    worksheet.write(row, 10, x.html)
                    worksheet.write(row, 11, ids[d])
                    worksheet.write(row, 12, dates[d].isoformat())
                    row += 1
                    num += 1
        
    workbook.close()
           
    wb.close()
    wb2.close()
    
    
    
									
              